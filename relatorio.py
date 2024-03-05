from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import web_scraping_goodreads as wsg

df_world = wsg.mais_lidos_da_semana_no_mundo()
df_brazil = wsg.mais_lidos_da_semana_no_brasil()
file_name = wsg.pivot_table_brasil_mundo(df_world, df_brazil)

wb = load_workbook(file_name)
sheet = wb['graficos']
barchart = BarChart()

data = Reference(sheet,
                 min_col=wb.active.min_column+1, max_col=wb.active.max_column,
                 min_row=wb.active.min_row, max_row=wb.active.max_row-1)
categories = Reference(sheet,
                 min_col=wb.active.min_column, max_col=wb.active.min_column,
                 min_row=wb.active.min_row+1, max_row=wb.active.max_row-1)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

sheet.add_chart(barchart, "B30")

barchart.title = "Leitores Brasileiros"
barchart.style = 18

sheet2 = wb['pivot_table']
barchart2 = BarChart()

data2 = Reference(sheet2,
                 min_col=sheet2.min_column+1, max_col=sheet2.max_column,
                 min_row=sheet2.min_row, max_row=sheet2.max_row)
categories2 = Reference(sheet2,
                 min_col=sheet2.min_column, max_col=sheet2.min_column,
                 min_row=sheet2.min_row+1, max_row=sheet2.max_row)

barchart2.add_data(data2, titles_from_data=True)
barchart2.set_categories(categories2)

sheet.add_chart(barchart2, "B10")

barchart2.title = "Livros mais lidos no Mundo no goodreads"
barchart2.style = 18

wb.save(file_name)